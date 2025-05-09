# -*- coding: utf-8 -*-
# @Time    : 2025/3/26 00:00
# @File    : 批量新建文件夹.py
# @Copyright : 破瓶子 (www.popingzi.com)
__author__ = "破瓶子"
__copyright__ = "Copyright (c) 2025 POPINGZI.COM (破瓶子)"

import re
import os
import sys
import json
import webbrowser
import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# 导入文件languages.py文件里定义的语言字典
from languages import LANGUAGES

# 语言配置文件
LANGUAGE_CONFIG_FILE = 'language_config.json'

# 加载语言设置
def load_language_setting():
    if os.path.exists(LANGUAGE_CONFIG_FILE):
        try:
            with open(LANGUAGE_CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f).get('language', 'zh')
        except Exception as e:
            print(f"加载语言设置时出错: {e}")
    return 'zh'

# 保存语言设置
def save_language_setting(lang):
    try:
        with open(LANGUAGE_CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump({'language': lang}, f)
    except Exception as e:
        print(f"保存语言设置时出错: {e}")

# 当前语言
current_language = load_language_setting()

# 全局变量，用于存储时间格式提示框
time_format_tip = None

# 全局变量，用于存储子文件夹格式提示框
folder_name_tip = None

# 全局变量，用于存储预设
presets = {}
PRESETS_FILE = 'presets.json'


def load_presets():
    if os.path.exists(PRESETS_FILE):
        try:
            with open(PRESETS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载预设时出错: {e}")
    return {}

def save_presets():
    try:
        with open(PRESETS_FILE, 'w', encoding='utf-8') as f:
            json.dump(presets, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print(f"保存预设时出错: {e}")

def is_valid_folder_name(folder_name):
    folder_name = folder_name.strip()  # 去除首尾空白字符
    invalid_chars = r'[\\/:*?"<>|]'  # 定义文件夹名称中不允许的字符
    if re.search(invalid_chars, folder_name):
        print(f"非法字符: {re.search(invalid_chars, folder_name).group()}")  # 调试信息
        return False
    return True

def create_folders(prefix, suffix, names, auto_increment, save_path):
    success_count = 0
    has_error = False  # 标志位，记录是否发生错误

    # 先检查前缀和后缀的合法性
    if not is_valid_folder_name(prefix):
        messagebox.showerror("错误", LANGUAGES[current_language]['error_invalid_prefix'].format(prefix=prefix))
        return
    if not is_valid_folder_name(suffix):
        messagebox.showerror("错误", LANGUAGES[current_language]['error_invalid_suffix'].format(suffix=suffix))
        return
    
    for name in names:
        if '├──' in name or '──' in name or '│   ' in name or '└──' in name:
            messagebox.showerror("错误", LANGUAGES[current_language]['error_invalid_characters'])
            has_error = True
            return



    try:
        for i, name in enumerate(names):
            sub_folders = name.split('>')
            current_path = save_path
            for j, sub_folder in enumerate(sub_folders):
                if auto_increment and j == 0:  # 只在一级目录添加自增数字
                    folder_name = f"{prefix}{sub_folder}{i + 1}{suffix}"  # 将自增数字添加到一级目录名称后面
                else:
                    folder_name = f"{prefix}{sub_folder}{suffix}"

                # 检查文件夹名称是否包含非法字符
                if not is_valid_folder_name(folder_name):
                    messagebox.showerror("错误", LANGUAGES[current_language]['error_invalid_folder_name'].format(folder_name=folder_name))
                    has_error = True  # 标记为发生错误
                    break

                # 拼接完整路径
                current_path = os.path.join(current_path, folder_name)
                if not os.path.exists(current_path):
                    os.makedirs(current_path)
                    if j == len(sub_folders) - 1:
                        success_count += 1
                else:
                    print(f"文件夹 {current_path} 已存在")  # 控制台输出，不弹窗
    except Exception as e:
        messagebox.showerror("错误", LANGUAGES[current_language]['error_create_folder'].format(error=e))
        has_error = True  # 标记为发生错误

    # 如果没有发生错误，再根据成功数量弹窗
    if not has_error:
        if success_count == len(names):
            messagebox.showinfo("成功", LANGUAGES[current_language]['success_all'])
        else:
            messagebox.showinfo("部分成功", LANGUAGES[current_language]['success_partial'].format(count=success_count))

def on_submit():
    prefix = prefix_entry.get()
    suffix = suffix_entry.get()
    names_text_content = names_text.get("1.0", tk.END).strip()
    # 支持按行和|分割文件夹名称
    names = []
    for line in names_text_content.splitlines():
        names.extend(line.split('|'))
    names = [name.strip() for name in names if name.strip()]
    auto_increment = auto_increment_var.get()  # 获取复选框状态
    save_path = save_path_entry.get()  # 获取用户选择的保存路径

    if not save_path:
        messagebox.showerror("错误", LANGUAGES[current_language]['error_no_save_path'])
        return

    if prefix:
        try:
            prefix = datetime.now().strftime(prefix)
        except ValueError:
            pass
    if suffix:
        try:
            suffix = datetime.now().strftime(suffix)
        except ValueError:
            pass

    create_folders(prefix, suffix, names, auto_increment, save_path)


def update_folder_count(event=None):
    names_text_content = names_text.get("1.0", tk.END).strip()
    names = []
    for line in names_text_content.splitlines():
        # 按行分割，并处理每行中的 "|" 和 ">"
        sub_names = line.split('|')
        for sub_name in sub_names:
            if '>' in sub_name:
                # 如果包含 ">"，说明有子文件夹，按 ">" 分割并累加所有层级的文件夹
                names.extend(sub_name.split('>'))
            else:
                # 否则直接添加为父文件夹
                names.append(sub_name.strip())
    names = [name.strip() for name in names if name.strip()]  # 去除空白项
    folder_count = len(names)
    folder_count_label.config(text=LANGUAGES[current_language]['folder_count_label'] + str(folder_count))


def select_save_path():
    # 打开文件夹选择对话框
    selected_path = filedialog.askdirectory()
    if selected_path:
        save_path_entry.delete(0, tk.END)
        save_path_entry.insert(0, selected_path)


# 显示时间格式提示框
def show_time_format_tip(event):
    global time_format_tip
    if time_format_tip is None or not tk.Toplevel.winfo_exists(time_format_tip):
        time_format_tip = tk.Toplevel(root)
        time_format_tip.wm_overrideredirect(True)
        time_format_tip.wm_geometry(f"+{event.x_root + 10}+{event.y_root + 10}")
        label = ttk.Label(time_format_tip, text=LANGUAGES[current_language]['time_format_tip'])
        label.pack(padx=5, pady=5)
        time_format_tip.configure(borderwidth=1, relief="solid")  # 添加边框


# 隐藏时间格式提示框
def hide_time_format_tip(event):
    global time_format_tip
    if time_format_tip is not None and tk.Toplevel.winfo_exists(time_format_tip):
        time_format_tip.destroy()

# 从 Excel 中导入
def import_from_excel():
    # 打开文件选择对话框
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        try:
            # 读取 Excel 文件
            df = pd.read_excel(file_path)
            if df.empty:
                messagebox.showerror(LANGUAGES[current_language]['error'], LANGUAGES[current_language]['empty_excel'])
                return
            # 转换为 > 分割的格式
            folder_names = []
            for index, row in df.iterrows():
                valid_folders = [str(folder) for folder in row if pd.notna(folder)]
                if valid_folders:
                    folder_names.append('>'.join(valid_folders))
            # 清空文本框并插入新内容
            names_text.delete("1.0", tk.END)
            names_text.insert(tk.END, '\n'.join(folder_names))
            messagebox.showinfo("成功", LANGUAGES[current_language]['import_success'])
        except Exception as e:
            messagebox.showerror(LANGUAGES[current_language]['error'], str(e))

# 导入Excel
def import_from_excel():
    # 打开文件选择对话框
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        try:
            # 读取 Excel 文件
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active

            # 获取合并单元格的信息
            merged_ranges = ws.merged_cells.ranges

            # 取消合并单元格并填充值
            for merged_range in list(merged_ranges):
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left_cell_value = ws.cell(row=min_row, column=min_col).value
                ws.unmerge_cells(range_string=merged_range.coord)
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        ws.cell(row=row, column=col, value=top_left_cell_value)

            # 转换为 > 分割的格式
            folder_names = []
            for row in ws.iter_rows(values_only=True):
                valid_folders = [str(folder) for folder in row if folder is not None]
                if valid_folders:
                    folder_names.append('>'.join(valid_folders))

            # 清空文本框并插入新内容
            names_text.delete("1.0", tk.END)
            names_text.insert(tk.END, '\n'.join(folder_names))
            messagebox.showinfo("成功", LANGUAGES[current_language]['import_success'])
        except Exception as e:
            messagebox.showerror(LANGUAGES[current_language]['error'], str(e))
    update_folder_count() # 更新文件夹数量
    

# 导出到Excel
def export_to_excel():
    # 获取文本框内容
    text = names_text.get("1.0", tk.END).strip()
    if not text:
        messagebox.showerror(LANGUAGES[current_language]['error'], LANGUAGES[current_language]['empty_excel'])
        return
    
    # 处理 | 分割的内容
    all_lines = [part for line in text.split('\n') for part in (line.split('|') if '|' in line else [line])]
    
    data = [line.split('>') for line in all_lines]
    
    # 找到最大列数
    max_columns = max(len(row) for row in data)
    
    # 补全缺失的列
    for row in data:
        while len(row) < max_columns:
            row.append(None)
    
    # 创建 Workbook 对象
    wb = Workbook()
    ws = wb.active
    
    # 写入数据
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            if value is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)
    # 合并相同的同级文件夹单元格
    def merge_cells(col_idx):
        start_row = None
        end_row = None
        current_value = None
        for row_idx in range(1, len(data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                if start_row is None:
                    start_row = row_idx
                    current_value = cell.value
                elif cell.value != current_value:
                    end_row = row_idx - 1
                    if end_row >= start_row:
                        col_letter = get_column_letter(col_idx)
                        ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{end_row}')
                        ws[f'{col_letter}{start_row}'].alignment = Alignment(vertical='center')
                    start_row = row_idx
                    current_value = cell.value
            else:
                # 如果遇到空单元格，且之前有需要合并的单元格，则进行合并
                if start_row is not None:
                    end_row = row_idx - 1
                    if end_row >= start_row:
                        col_letter = get_column_letter(col_idx)
                        ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{end_row}')
                        ws[f'{col_letter}{start_row}'].alignment = Alignment(vertical='center')
                    start_row = None
                    current_value = None

        # 处理最后一组单元格
        if start_row is not None:
            # 找到最后一个有值的单元格的行号
            last_non_empty_row = len(data)
            while last_non_empty_row > 0 and ws.cell(row=last_non_empty_row, column=col_idx).value is None:
                last_non_empty_row -= 1
            end_row = last_non_empty_row
            if end_row >= start_row:
                col_letter = get_column_letter(col_idx)
                ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{end_row}')
                ws[f'{col_letter}{start_row}'].alignment = Alignment(vertical='center')
    
    for col_idx in range(1, max_columns + 1):
        merge_cells(col_idx)
    
    # 打开文件保存对话框
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx;*.xls")])
    if not file_path:
        return
    
    try:
        # 保存为 Excel 文件
        wb.save(file_path)
        messagebox.showinfo("成功", LANGUAGES[current_language]['export_success'])
    except PermissionError as e:
        messagebox.showerror(LANGUAGES[current_language]['error'], LANGUAGES[current_language]['permission_error'])
    except IOError as e:
        messagebox.showerror(LANGUAGES[current_language]['error'], LANGUAGES[current_language]['io_error'])
    except Exception as e:
        messagebox.showerror(LANGUAGES[current_language]['error'], str(e))
            
# 切换语言
def switch_language(lang):
    global current_language
    if lang == current_language:
        return  # 如果选择的语言和当前语言相同，直接返回，避免报错
    current_language = lang
    save_language_setting(lang)  # 保存语言设置
    root.title(LANGUAGES[current_language]['title'])  # 更新窗口标题
    save_path_label.config(text=LANGUAGES[current_language]['save_path_label'])  # 更新保存路径标签
    select_path_button.config(text=LANGUAGES[current_language]['select_path_button'])  # 更新选择路径按钮
    prefix_label.config(text=LANGUAGES[current_language]['prefix_label'])  # 更新前缀标签
    prefix_tip_label.config(text=LANGUAGES[current_language]['prefix_tip_label'])  # 更新前缀提示标签
    suffix_label.config(text=LANGUAGES[current_language]['suffix_label'])  # 更新后缀标签
    suffix_tip_label.config(text=LANGUAGES[current_language]['suffix_tip_label'])  # 更新后缀提示标签
    folder_name_label.config(text=LANGUAGES[current_language]['folder_name_label'])  # 更新文件夹名称标签
    folder_count_label.config(text=LANGUAGES[current_language]['folder_count_label'] + str(len(get_folder_names())))  # 更新文件夹数量标签
    auto_increment_check.config(text=LANGUAGES[current_language]['auto_increment_check'])  # 更新自动递增标签
    submit_button.config(text=LANGUAGES[current_language]['submit_button'])  # 更新提交按钮标签
    folder_name_tip_label.config(text=LANGUAGES[current_language]['folder_name_tip_title'])  # 更新文件夹名称提示标题


    # 重新创建菜单栏
    menubar = tk.Menu(root)
    lang_menu = tk.Menu(menubar, tearoff=0)
    lang_menu.add_command(label=LANGUAGES[current_language]['lang_zh'], command=lambda: switch_language('zh'))
    lang_menu.add_command(label=LANGUAGES[current_language]['lang_en'], command=lambda: switch_language('en'))
    lang_menu.add_command(label=LANGUAGES[current_language]['lang_ja'], command=lambda: switch_language('ja'))
    lang_menu.add_command(label=LANGUAGES[current_language]['lang_ko'], command=lambda: switch_language('ko'))
    menubar.add_cascade(label=LANGUAGES[current_language]['lang_menu'], menu=lang_menu)

    # 添加预设菜单
    preset_menu = tk.Menu(menubar, tearoff=0)
    preset_menu.add_command(label=LANGUAGES[current_language]['save_preset'], command=save_preset)
    for name in presets.keys():
        preset_menu.add_command(label=name, command=lambda n=name: load_preset(n))
    preset_menu.add_command(label=LANGUAGES[current_language]['delete_preset'], command=delete_preset)
    menubar.add_cascade(label=LANGUAGES[current_language]['preset_menu'], menu=preset_menu)

    # 添加文件菜单
    file_menu = tk.Menu(menubar, tearoff=0)
    file_menu.add_command(label=LANGUAGES[current_language]['get_all_folders'], command=get_all_folders)
    file_menu.add_command(label=LANGUAGES[current_language]['get_all_files'], command=get_all_files)
    file_menu.add_command(label=LANGUAGES[current_language]['get_all_folders_and'], command=get_all_folders_and_subfolders)
    file_menu.add_command(label=LANGUAGES[current_language]['get_all_folders_and_subfolders'], command=get_all_folders_files_and_subfolders)
    menubar.add_cascade(label=LANGUAGES[current_language]['file_menu'], menu=file_menu)

    edit_menu = tk.Menu(menubar, tearoff=0)
    edit_menu.add_command(label=LANGUAGES[current_language]['remove_duplicates'], command=remove_duplicates)
    edit_menu.add_command(label=LANGUAGES[current_language]['merge_and_sort'], command=merge_and_sort)
    edit_menu.add_command(label=LANGUAGES[current_language]['sort_ascending'], command=sort_ascending)
    edit_menu.add_command(label=LANGUAGES[current_language]['sort_descending'], command=sort_descending)
    edit_menu.add_command(label=LANGUAGES[current_language]['convert_to_structure_diagram'], command=convert_to_structure_diagram)
    menubar.add_cascade(label=LANGUAGES[current_language]['edit_menu'], menu=edit_menu)

    # 添加导入/导出功能
    export_menu = tk.Menu(menubar, tearoff=0)
    export_menu.add_command(label=LANGUAGES[current_language]['from_excel'], command=import_from_excel) #导入Excel
    export_menu.add_command(label=LANGUAGES[current_language]['to_excel'], command=export_to_excel) #导出到Excel
    menubar.add_cascade(label=LANGUAGES[current_language]['export_menu'], menu=export_menu)

    help_menu = tk.Menu(menubar, tearoff=0)
    help_menu.add_command(label=LANGUAGES[current_language]['help_shortcut'], command=help_shortcut) # 快捷键帮助信息
    help_menu.add_command(label=LANGUAGES[current_language]['help_admin'], command=show_admin) # 软件作者信息
    help_menu.add_command(label=LANGUAGES[current_language]['help_url'], command=help_url) # 作者主页
    help_menu.add_command(label=LANGUAGES[current_language]['excel_tip'], command=excel_tip) # 表格格式要求
    menubar.add_cascade(label=LANGUAGES[current_language]['help_menu'], menu=help_menu)

    root.config(menu=menubar)


# 获取文件夹名称列表
def get_folder_names():
    names_text_content = names_text.get("1.0", tk.END).strip()
    if '|' in names_text_content:
        names = names_text_content.split('|')
    elif '\n' in names_text_content:
        names = names_text_content.split('\n')
    elif '>' in names_text_content:
        names = names_text_content.split('>')
    else:
        names = names_text_content.split('\n')
    
    return [name for name in names if name]


# 保存预设
def save_preset():
    def save():
        preset_name = entry.get()
        if preset_name:
            prefix = prefix_entry.get()
            suffix = suffix_entry.get()
            auto_increment = auto_increment_var.get()
            save_path = save_path_entry.get()
            presets[preset_name] = {
                'prefix': prefix,
                'suffix': suffix,
                'auto_increment': auto_increment,
                'save_path': save_path
            }
            save_presets() 
            # 更新预设菜单
            update_preset_menu()
            update_menu_language()
            top.destroy()
        else:
            messagebox.showerror(LANGUAGES[current_language]['error'], LANGUAGES[current_language]['preset_name_empty'])

    top = tk.Toplevel(root)
    top.title(LANGUAGES[current_language]['save_preset'])
    label = ttk.Label(top, text=LANGUAGES[current_language]['preset_name'])
    label.pack(padx=10, pady=5)
    entry = ttk.Entry(top)
    entry.pack(padx=10, pady=5)
    button = ttk.Button(top, text=LANGUAGES[current_language]['save'], command=save)
    button.pack(padx=10, pady=10)


# 加载预设
def load_preset(preset_name):
    if preset_name not in presets:# 如果预设不存在，则显示错误消息并返回
        messagebox.showerror(LANGUAGES[current_language]['error'], LANGUAGES[current_language]['preset_not_found'])
        return
    preset = presets[preset_name]
    prefix_entry.delete(0, tk.END)
    prefix_entry.insert(0, preset['prefix'])
    suffix_entry.delete(0, tk.END)
    suffix_entry.insert(0, preset['suffix'])
    auto_increment_var.set(preset['auto_increment'])
    save_path_entry.delete(0, tk.END)
    save_path_entry.insert(0, preset['save_path'])


# 删除预设
def delete_preset():
    def confirm_delete():
        preset_name = var.get()
        if preset_name in presets:
            del presets[preset_name]
            save_presets() 
            # 更新预设菜单
            update_preset_menu()
            update_menu_language()
            top.destroy()
        else:
            messagebox.showerror(LANGUAGES[current_language]['error'], LANGUAGES[current_language]['preset_not_found'])

    top = tk.Toplevel(root)
    top.title(LANGUAGES[current_language]['delete_preset'])
    var = tk.StringVar()
    var.set(list(presets.keys())[0] if presets else "")
    option_menu = ttk.OptionMenu(top, var, *presets.keys())
    option_menu.pack(padx=10, pady=5)
    button = ttk.Button(top, text=LANGUAGES[current_language]['delete'], command=confirm_delete)
    button.pack(padx=10, pady=10)


# 更新菜单语言显示
def update_menu_language():
    menubar = tk.Menu(root)
    lang_menu = tk.Menu(menubar, tearoff=0)
    lang_menu.add_command(label=LANGUAGES[current_language]['lang_zh'], command=lambda: switch_language('zh'))
    lang_menu.add_command(label=LANGUAGES[current_language]['lang_en'], command=lambda: switch_language('en'))
    lang_menu.add_command(label=LANGUAGES[current_language]['lang_ja'], command=lambda: switch_language('ja'))
    lang_menu.add_command(label=LANGUAGES[current_language]['lang_ko'], command=lambda: switch_language('ko'))
    menubar.add_cascade(label=LANGUAGES[current_language]['lang_menu'], menu=lang_menu)

    # 添加预设菜单
    preset_menu = tk.Menu(menubar, tearoff=0)
    preset_menu.add_command(label=LANGUAGES[current_language]['save_preset'], command=save_preset)
    for name in presets.keys():
        preset_menu.add_command(label=name, command=lambda n=name: load_preset(n))
    preset_menu.add_command(label=LANGUAGES[current_language]['delete_preset'], command=delete_preset)
    menubar.add_cascade(label=LANGUAGES[current_language]['preset_menu'], menu=preset_menu)

    # 添加文件菜单
    file_menu = tk.Menu(menubar, tearoff=0)
    file_menu.add_command(label=LANGUAGES[current_language]['get_all_folders'], command=get_all_folders)
    file_menu.add_command(label=LANGUAGES[current_language]['get_all_files'], command=get_all_files)
    file_menu.add_command(label=LANGUAGES[current_language]['get_all_folders_and'], command=get_all_folders_and_subfolders)
    file_menu.add_command(label=LANGUAGES[current_language]['get_all_folders_and_subfolders'], command=get_all_folders_files_and_subfolders)
    menubar.add_cascade(label=LANGUAGES[current_language]['file_menu'], menu=file_menu)

    edit_menu = tk.Menu(menubar, tearoff=0)
    edit_menu.add_command(label=LANGUAGES[current_language]['remove_duplicates'], command=remove_duplicates)
    edit_menu.add_command(label=LANGUAGES[current_language]['merge_and_sort'], command=merge_and_sort)
    edit_menu.add_command(label=LANGUAGES[current_language]['sort_ascending'], command=sort_ascending)
    edit_menu.add_command(label=LANGUAGES[current_language]['sort_descending'], command=sort_descending)
    edit_menu.add_command(label=LANGUAGES[current_language]['convert_to_structure_diagram'], command=convert_to_structure_diagram)
    menubar.add_cascade(label=LANGUAGES[current_language]['edit_menu'], menu=edit_menu)

    # 添加导入/导出功能
    export_menu = tk.Menu(menubar, tearoff=0)
    export_menu.add_command(label=LANGUAGES[current_language]['from_excel'], command=import_from_excel) #导入Excel
    export_menu.add_command(label=LANGUAGES[current_language]['to_excel'], command=export_to_excel) #导出到Excel
    menubar.add_cascade(label=LANGUAGES[current_language]['export_menu'], menu=export_menu)

    help_menu = tk.Menu(menubar, tearoff=0)
    help_menu.add_command(label=LANGUAGES[current_language]['help_shortcut'], command=help_shortcut) #快捷键帮助信息
    help_menu.add_command(label=LANGUAGES[current_language]['help_admin'], command=show_admin) #软件作者信息
    help_menu.add_command(label=LANGUAGES[current_language]['help_url'], command=help_url) #作者主页
    help_menu.add_command(label=LANGUAGES[current_language]['excel_tip'], command=excel_tip)
    menubar.add_cascade(label=LANGUAGES[current_language]['help_menu'], menu=help_menu)

    root.config(menu=menubar)

# 获取所有文件夹及文件名称
def get_all_folders_files_and_subfolders():
    path = save_path_entry.get()
    if not path or not os.path.exists(path):
        messagebox.showerror(LANGUAGES[current_language]['error'],
                             LANGUAGES[current_language]['invalid_path'])
        return
    items = []
    for root_dir, dirs, files in os.walk(path):
        relative_path = os.path.relpath(root_dir, path)
        if relative_path == ".":
            relative_path = ""
        for dir in dirs:
            if relative_path:
                folder_structure = f"{relative_path.replace(os.sep, '>')}>{dir}"
            else:
                folder_structure = dir
            items.append(folder_structure)
        for file in files:
            if relative_path:
                file_structure = f"{relative_path.replace(os.sep, '>')}>{file}"
            else:
                file_structure = file
            items.append(file_structure)
    names_text.delete("1.0", tk.END)
    names_text.insert(tk.END, "\n".join(items))
    update_folder_count()


# 获取指定路径下的所有文件夹及子文件夹名称
def get_all_folders_and_subfolders():
    path = save_path_entry.get()
    if os.path.exists(path):
        folder_paths = []
        def traverse_folders(current_path, parent_folder_names=[]):
            for item in os.listdir(current_path):
                item_path = os.path.join(current_path, item)
                if os.path.isdir(item_path):
                    new_parent_folder_names = parent_folder_names + [item]
                    folder_paths.append('>'.join(new_parent_folder_names))
                    traverse_folders(item_path, new_parent_folder_names)
        traverse_folders(path)
        names_text.delete(1.0, tk.END)
        for folder_path in folder_paths:
            names_text.insert(tk.END, folder_path + '\n')
        update_folder_count()
    else:
        messagebox.showerror(LANGUAGES[current_language]['error'], LANGUAGES[current_language]['invalid_path'])


# 更新预设菜单
def update_preset_menu():
    preset_menu.delete(0, tk.END)
    preset_menu.add_command(label=LANGUAGES[current_language]['save_preset'], command=save_preset)
    for name in presets.keys():
        preset_menu.add_command(label=name, command=lambda n=name: load_preset(n))
    preset_menu.add_command(label=LANGUAGES[current_language]['delete_preset'], command=delete_preset)

# 获取指定路径下的所有文件夹名称
def get_all_folders():
    path = save_path_entry.get()
    if os.path.exists(path):
        folders = [name for name in os.listdir(path) if os.path.isdir(os.path.join(path, name))]
        names_text.delete(1.0, tk.END)
        for folder in folders:
            names_text.insert(tk.END, folder + '\n')
        update_folder_count()
    else:
        messagebox.showerror(LANGUAGES[current_language]['error'], LANGUAGES[current_language]['invalid_path'])

# 获取指定路径下的所有文件名称
def get_all_files():
    path = save_path_entry.get()
    if os.path.exists(path):
        files = [name for name in os.listdir(path) if os.path.isfile(os.path.join(path, name))]
        names_text.delete(1.0, tk.END)
        for file in files:
            names_text.insert(tk.END, file + '\n')
        update_folder_count()
    else:
        messagebox.showerror(LANGUAGES[current_language]['error'], LANGUAGES[current_language]['invalid_path'])

# 去除重复的一级文件夹名
def remove_duplicates():
    text = names_text.get("1.0", tk.END)
    # 先按换行符分割，再按 | 分割
    all_items = []
    for line in text.splitlines():
        all_items.extend(line.split('|'))
    unique_items = list(set(all_items))
    new_text = '\n'.join(unique_items)
    names_text.delete("1.0", tk.END)
    names_text.insert(tk.END, new_text)
    update_folder_count()

# 合并整理文件夹名
def merge_and_sort():
    text = names_text.get("1.0", tk.END)
    all_items = []
    for line in text.splitlines():
        all_items.extend(line.split('|'))
    folder_structure = {}
    for item in all_items:
        parts = item.rstrip('>').split('>')
        current_level = folder_structure
        for part in parts:
            if part not in current_level:
                current_level[part] = {}
            current_level = current_level[part]

    def flatten_structure(structure, prefix=""):
        result = []
        for key, value in structure.items():
            new_prefix = f"{prefix}>{key}" if prefix else key
            if value:
                result.extend(flatten_structure(value, new_prefix))
            else:
                result.append(new_prefix)
        return result

    sorted_lines = flatten_structure(folder_structure)
    new_text = '\n'.join(sorted_lines)
    names_text.delete("1.0", tk.END)
    names_text.insert(tk.END, new_text)
    update_folder_count()

# 升序排序函数
def sort_ascending():
    names_text_content = names_text.get("1.0", tk.END).strip()
    all_names = []
    for line in names_text_content.splitlines():
        all_names.extend(line.split('|'))
    all_names = [name.strip() for name in all_names if name.strip()]
    sorted_names = sorted(all_names)
    names_text.delete("1.0", tk.END)
    output = '|'.join(sorted_names)
    names_text.insert(tk.END, output)
    update_folder_count()

# 降序排序函数
def sort_descending():
    names_text_content = names_text.get("1.0", tk.END).strip()
    all_names = []
    for line in names_text_content.splitlines():
        all_names.extend(line.split('|'))
    all_names = [name.strip() for name in all_names if name.strip()]
    sorted_names = sorted(all_names, reverse=True)
    names_text.delete("1.0", tk.END)
    output = '|'.join(sorted_names)
    names_text.insert(tk.END, output)
    update_folder_count()

# 显示子文件夹格式提示框
def show_folder_name_tip(event):
    global folder_name_tip
    if folder_name_tip is None or not tk.Toplevel.winfo_exists(folder_name_tip):
        folder_name_tip = tk.Toplevel(root)
        folder_name_tip.wm_overrideredirect(True)
        folder_name_tip.wm_geometry(f"+{event.x_root + 10}+{event.y_root + 10}")
        label = ttk.Label(folder_name_tip, text=LANGUAGES[current_language]['folder_name_tip'])
        label.pack(padx=5, pady=5)
        folder_name_tip.configure(borderwidth=1, relief="solid")  # 添加边框

# 隐藏子文件夹格式提示框
def hide_folder_name_tip(event):
    global folder_name_tip
    if folder_name_tip is not None and tk.Toplevel.winfo_exists(folder_name_tip):
        folder_name_tip.destroy()

# 转换为结构图
def convert_to_structure_diagram():
    text = names_text.get("1.0", tk.END).strip()
    # 预处理：将 | 分隔的部分拆分成多行
    text = text.replace('|', '\n')
    lines = text.split('\n')
    structure = {}
    for line in lines:
        parts = line.split('>')
        current_level = structure
        for part in parts:
            part = part.strip()
            if part not in current_level:
                current_level[part] = {}
            current_level = current_level[part]

    def print_structure(level, indent=0):
        result = []
        keys = list(level.keys())
        for i, key in enumerate(keys):
            if indent == 0:
                result.append(key)
            else:
                if i == len(keys) - 1:
                    # 如果是最后一个子项，使用 └──
                    result.append(' ' + '│   ' * (indent - 1) + '└── ' + key)
                else:
                    # 否则使用 ├──
                    result.append(' ' + '│   ' * (indent - 1) + '├── ' + key)
            if level[key]:
                result.extend(print_structure(level[key], indent + 1))
        return result

    diagrams = []
    for key, value in structure.items():
        diagram_lines = [key] + print_structure(value, 1)
        diagrams.append('\n'.join(diagram_lines))

    final_diagram = '\n\n'.join(diagrams)
    names_text.delete("1.0", tk.END)
    names_text.insert("1.0", final_diagram)

# 添加帮助菜单相关函数
def help_shortcut():
    messagebox.showinfo(LANGUAGES[current_language]['help_shortcut'], LANGUAGES[current_language]['shortcut_ctrl'])

def show_admin():
    messagebox.showinfo(LANGUAGES[current_language]['help_admin'],LANGUAGES[current_language]['admin_info'])
def help_url():
    # 在浏览器中打开网址
    webbrowser.open("https://www.popingzi.com")

def excel_tip():
    messagebox.showinfo(LANGUAGES[current_language]['excel_tip'], LANGUAGES[current_language]['help_excel'])



# 创建主窗口
root = tk.Tk()
root.title(LANGUAGES[current_language]['title'])
root.geometry("600x600")  # 调整窗口大小
root.minsize(600, 500)  # 最小窗口大小限制

# 判断是否是打包后的程序
if getattr(sys, 'frozen', False):
    # 打包后的程序路径
    base_dir = sys._MEIPASS
else:
    # 开发时的程序路径
    base_dir = os.path.dirname(os.path.abspath(__file__))

icon_path = os.path.join(base_dir, "new.ico")

# 检查图标文件是否存在
if os.path.exists(icon_path):
    try:
        root.iconbitmap(icon_path)
    except Exception as e:
        print(f"设置图标时出错: {e}")
else:
    print("图标文件未找到。")

# 使用 ttk 美化样式
style = ttk.Style()
style.configure("TLabel", font=("微软雅黑", 10))
style.configure("TButton", font=("微软雅黑", 10), padding=5)
style.configure("TEntry", font=("微软雅黑", 10), padding=5)
style.configure("TText", font=("微软雅黑", 10), padding=5)

# 创建菜单栏
menubar = tk.Menu(root)
lang_menu = tk.Menu(menubar, tearoff=0)
lang_menu.add_command(label=LANGUAGES[current_language]['lang_zh'], command=lambda: switch_language('zh'))
lang_menu.add_command(label=LANGUAGES[current_language]['lang_en'], command=lambda: switch_language('en'))
lang_menu.add_command(label=LANGUAGES[current_language]['lang_ja'], command=lambda: switch_language('ja'))
lang_menu.add_command(label=LANGUAGES[current_language]['lang_ko'], command=lambda: switch_language('ko'))
menubar.add_cascade(label=LANGUAGES[current_language]['lang_menu'], menu=lang_menu)

# 添加预设菜单
preset_menu = tk.Menu(menubar, tearoff=0)
preset_menu.add_command(label=LANGUAGES[current_language]['save_preset'], command=save_preset)

presets = load_presets()
for name in presets.keys():
    preset_menu.add_command(label=name, command=lambda n=name: load_preset(n))
preset_menu.add_command(label=LANGUAGES[current_language]['delete_preset'], command=delete_preset)
menubar.add_cascade(label=LANGUAGES[current_language]['preset_menu'], menu=preset_menu)

# 添加获取文件夹和文件名称的菜单
file_menu = tk.Menu(menubar, tearoff=0)
file_menu.add_command(label=LANGUAGES[current_language]['get_all_folders'], command=get_all_folders)
file_menu.add_command(label=LANGUAGES[current_language]['get_all_files'], command=get_all_files)
file_menu.add_command(label=LANGUAGES[current_language]['get_all_folders_and'], command=get_all_folders_and_subfolders)
file_menu.add_command(label=LANGUAGES[current_language]['get_all_folders_and_subfolders'], command=get_all_folders_files_and_subfolders)
menubar.add_cascade(label=LANGUAGES[current_language]['file_menu'], menu=file_menu)

# 去重功能
edit_menu = tk.Menu(menubar, tearoff=0)
edit_menu.add_command(label=LANGUAGES[current_language]['remove_duplicates'], command=remove_duplicates)
edit_menu.add_command(label=LANGUAGES[current_language]['merge_and_sort'], command=merge_and_sort)
edit_menu.add_command(label=LANGUAGES[current_language]['sort_ascending'], command=sort_ascending)
edit_menu.add_command(label=LANGUAGES[current_language]['sort_descending'], command=sort_descending)
edit_menu.add_command(label=LANGUAGES[current_language]['convert_to_structure_diagram'], command=convert_to_structure_diagram)
menubar.add_cascade(label=LANGUAGES[current_language]['edit_menu'], menu=edit_menu)

# 添加导入/导出功能
export_menu = tk.Menu(menubar, tearoff=0)
export_menu.add_command(label=LANGUAGES[current_language]['from_excel'], command=import_from_excel) #导入Excel
export_menu.add_command(label=LANGUAGES[current_language]['to_excel'], command=export_to_excel) #导出到Excel
menubar.add_cascade(label=LANGUAGES[current_language]['export_menu'], menu=export_menu)

# 添加帮助菜单
help_menu = tk.Menu(menubar, tearoff=0)
help_menu.add_command(label=LANGUAGES[current_language]['help_shortcut'], command=help_shortcut) #快捷键帮助信息
help_menu.add_command(label=LANGUAGES[current_language]['help_admin'], command=show_admin) #软件作者信息
help_menu.add_command(label=LANGUAGES[current_language]['help_url'], command=help_url) #作者主页
help_menu.add_command(label=LANGUAGES[current_language]['excel_tip'], command=excel_tip)
menubar.add_cascade(label=LANGUAGES[current_language]['help_menu'], menu=help_menu)

root.config(menu=menubar)


# 使用网格布局管理器，设置列权重
root.columnconfigure(0, weight=1)  # 让第一列占据全部宽度

# 创建文件夹保存路径选择框
save_path_label = ttk.Label(root, text=LANGUAGES[current_language]['save_path_label'])
save_path_label.grid(row=0, column=0, padx=10, pady=5, sticky="w")
save_path_frame = ttk.Frame(root)
save_path_frame.grid(row=1, column=0, padx=10, pady=5, sticky="ew")

save_path_entry = ttk.Entry(save_path_frame)
save_path_entry.pack(side="left", fill="x", expand=True)  # 左侧输入框

select_path_button = ttk.Button(save_path_frame, text=LANGUAGES[current_language]['select_path_button'], command=select_save_path)
select_path_button.pack(side="right", padx=(10, 0))  # 右侧按钮

# 创建输入框和标签（上下布局）
prefix_label_frame = ttk.Frame(root)
prefix_label_frame.grid(row=2, column=0, padx=10, pady=5, sticky="w")
prefix_label = ttk.Label(prefix_label_frame, text=LANGUAGES[current_language]['prefix_label'])
prefix_label.pack(side="left")
prefix_tip_label = ttk.Label(prefix_label_frame, text=LANGUAGES[current_language]['prefix_tip_label'], foreground="red", cursor="hand2")
prefix_tip_label.pack(side="left")
prefix_tip_label.bind("<Enter>", show_time_format_tip)
prefix_tip_label.bind("<Leave>", hide_time_format_tip)

prefix_entry = ttk.Entry(root)
prefix_entry.grid(row=3, column=0, padx=10, pady=5, sticky="ew")  # 宽度与窗口一致

suffix_label_frame = ttk.Frame(root)
suffix_label_frame.grid(row=4, column=0, padx=10, pady=5, sticky="w")
suffix_label = ttk.Label(suffix_label_frame, text=LANGUAGES[current_language]['suffix_label'])
suffix_label.pack(side="left")
suffix_tip_label = ttk.Label(suffix_label_frame, text=LANGUAGES[current_language]['suffix_tip_label'], foreground="red", cursor="hand2")
suffix_tip_label.pack(side="left")
suffix_tip_label.bind("<Enter>", show_time_format_tip)
suffix_tip_label.bind("<Leave>", hide_time_format_tip)

suffix_entry = ttk.Entry(root)
suffix_entry.grid(row=5, column=0, padx=10, pady=5, sticky="ew")  # 宽度与窗口一致

# 创建文件夹名称标签和文件夹数量标签（在同一行）
folder_name_frame = ttk.Frame(root)  # 修改为 ttk.Frame
folder_name_frame.grid(row=6, column=0, padx=10, pady=5, sticky="ew")

folder_name_label = ttk.Label(folder_name_frame, text=LANGUAGES[current_language]['folder_name_label'])
folder_name_label.pack(side="left")  # 左侧
folder_count_label = ttk.Label(folder_name_frame, text=LANGUAGES[current_language]['folder_count_label'] + "0")
folder_count_label.pack(side="right")  # 右侧

# 添加子文件夹创建帮助提示标签
folder_name_tip_label = ttk.Label(folder_name_frame, text=LANGUAGES[current_language]['folder_name_tip_title'], foreground="red", cursor="hand2")
folder_name_tip_label.pack(side="left", padx=(5, 0))
folder_name_tip_label.bind("<Enter>", show_folder_name_tip)
folder_name_tip_label.bind("<Leave>", hide_folder_name_tip)

# 创建多行文本框
names_text = tk.Text(root, height=5, font=("微软雅黑", 10),undo=True)
names_text.grid(row=7, column=0, padx=10, pady=5, sticky="nsew")  # 可拉伸，宽度与窗口一致

# 定义一个函数，在按下回车或输入|时设置分隔符
def set_undo_separator(event):
    names_text.edit_separator()
    return None

# 绑定回撤和重做快捷键
names_text.bind("<Control-z>", lambda event: names_text.edit_undo())
names_text.bind("<Control-y>", lambda event: names_text.edit_redo())

# 绑定回车和|事件到设置分隔符的函数
names_text.bind("<Return>", set_undo_separator)
names_text.bind("|", set_undo_separator)

# 绑定事件监听器，实时更新文件夹数量
names_text.bind("<KeyRelease>", update_folder_count)

# 添加自增数字复选框
auto_increment_var = tk.BooleanVar()  # 用于存储复选框状态
auto_increment_check = ttk.Checkbutton(root, text=LANGUAGES[current_language]['auto_increment_check'], variable=auto_increment_var)
auto_increment_check.grid(row=8, column=0, padx=10, pady=5, sticky="w")

# 创建提交按钮
submit_button = ttk.Button(root, text=LANGUAGES[current_language]['submit_button'], command=on_submit)
submit_button.grid(row=9, column=0, padx=10, pady=10, sticky="ew")  # 宽度与窗口一致

# 设置行权重，使多行文本框可以拉伸
root.rowconfigure(7, weight=1)

# 运行主循环
root.mainloop()